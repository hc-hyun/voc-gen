from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def _human_review_metrics(path: Path | None) -> dict:
    if path is None or not path.exists():
        return {
            "status": "NOT_PROVIDED",
            "completed_rows": 0,
            "label_match_rate": None,
            "naturalness_average": None,
            "channel_fit_average": None,
            "cause_overclaim_rate": None,
        }
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["human_review"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    completed = [
        row
        for row in rows
        if row[12] in {"YES", "NO"}
        and isinstance(row[13], (int, float))
        and isinstance(row[14], (int, float))
        and row[15] in {"YES", "NO"}
    ]
    if not completed:
        return {
            "status": "PENDING",
            "completed_rows": 0,
            "label_match_rate": None,
            "naturalness_average": None,
            "channel_fit_average": None,
            "cause_overclaim_rate": None,
        }
    metrics = {
        "status": "COMPLETE" if len(completed) == len(rows) else "PARTIAL",
        "completed_rows": len(completed),
        "label_match_rate": sum(row[12] == "YES" for row in completed)
        / len(completed),
        "naturalness_average": sum(float(row[13]) for row in completed)
        / len(completed),
        "channel_fit_average": sum(float(row[14]) for row in completed)
        / len(completed),
        "cause_overclaim_rate": sum(row[15] == "YES" for row in completed)
        / len(completed),
    }
    return metrics


def build_release_reports(
    manifest_path: Path,
    validation_path: Path,
    output_dir: Path,
    version: str,
    human_review_path: Path | None = None,
) -> dict:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    validation = json.loads(validation_path.read_text(encoding="utf-8"))
    human = _human_review_metrics(human_review_path)
    deterministic_passed = bool(validation.get("passed"))
    human_passed = (
        human["status"] == "COMPLETE"
        and human["completed_rows"] >= 50
        and human["label_match_rate"] is not None
        and human["label_match_rate"] >= 0.95
        and human["naturalness_average"] is not None
        and human["naturalness_average"] >= 4.0
        and human["cause_overclaim_rate"] == 0
    )
    release_status = (
        "RELEASE_READY"
        if deterministic_passed and human_passed
        else "CANDIDATE_PENDING_HUMAN_REVIEW"
        if deterministic_passed
        else "FAILED_AUTOMATIC_VALIDATION"
    )
    distributions = validation.get("distributions", {})
    profile = manifest["profile"]
    output_dir.mkdir(parents=True, exist_ok=True)
    card_path = output_dir / f"DATASET_CARD_{version}.md"
    qa_path = output_dir / f"QA_REPORT_{version}.md"
    report_path = output_dir / f"release_{version}.json"

    card_path.write_text(
        f"""# Galaxy Synthetic Raw VoC Dataset {version}

## 상태

`{release_status}`

## 개요

- 문서 수: {manifest['row_count']:,}
- 기준 시나리오: Galaxy VoC Data Pack v0.1의 500개
- 생성기: `{manifest['generator_version']}`
- 생성 방식: clean text 후 profile별 표면 변형과 통제된 noise 적용
- 본문: `{manifest['data_file']}`
- 생성 이력: `{manifest['generation_file']}`
- 라이선스/용도: 내부 연구·정규화 파이프라인 개발용 합성 데이터

## 분포

```json
{json.dumps(distributions, ensure_ascii=False, indent=2)}
```

## 권장 사용

- `raw_text`만 정규화 모델 입력으로 사용한다.
- `issues[]`와 generation sidecar는 학습 라벨·오류 분석에만 사용한다.
- 같은 `parent_scenario_id`의 파생 문서는 원래 `dataset_split`을 유지한다.
- 합성 TEST 결과를 실제 운영 성능으로 해석하지 않는다.

## 알려진 한계

- 의미 기준점은 500개이므로 10만/100만 건은 표면 다양성 확대 데이터다.
- 영문은 ontology 기반 템플릿이며 전문 번역 데이터가 아니다.
- exact/normalized duplicate는 검사하지만 embedding 기반 semantic duplicate는
  별도 분석이 필요하다.
- 실제 고객 빈도 분포나 실제 개인정보를 반영하지 않는다.
- 사람 평가가 완료되지 않은 후보 데이터는 운영 릴리스로 간주하지 않는다.
""",
        encoding="utf-8",
    )
    checks = "\n".join(
        f"- [{'x' if check['passed'] else ' '}] `{check['name']}`: {check['actual']}"
        for check in validation["checks"]
    )
    qa_path.write_text(
        f"""# QA Report {version}

## 판정

- 릴리스 상태: **{release_status}**
- 결정적 검사: **{'통과' if deterministic_passed else '실패'}**
- 사람 검사: **{'통과' if human_passed else '미통과/대기'}**

## 결정적 검사

{checks}

- manifest 건수 일치: `{validation.get('row_count_matches_manifest')}`
- quarantine: `{validation.get('artifacts', {}).get('quarantine_file')}`
- validation results: `{validation.get('artifacts', {}).get('validation_results_file')}`

## 사람 평가

```json
{json.dumps(human, ensure_ascii=False, indent=2)}
```

합격 기준은 완료 표본 50건 이상, 라벨 일치율 95% 이상, 자연스러움 평균
4.0/5 이상, 원인 과장 0건이다.
""",
        encoding="utf-8",
    )
    release = {
        "version": version,
        "status": release_status,
        "manifest": str(manifest_path),
        "validation": str(validation_path),
        "dataset_card": str(card_path),
        "qa_report": str(qa_path),
        "deterministic_passed": deterministic_passed,
        "human_review": human,
        "profile": profile,
    }
    report_path.write_text(
        json.dumps(release, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return {
        "status": release_status,
        "dataset_card": card_path,
        "qa_report": qa_path,
        "release_report": report_path,
    }
