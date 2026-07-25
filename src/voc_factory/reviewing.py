from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

from .source import load_scenarios


def _read_jsonl(path: Path) -> list[dict]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line
    ]


def _select_review_rows(
    records: list[tuple[dict, dict]],
    scenario_by_id: dict,
    sample_size: int,
    min_per_stratum: int,
) -> list[tuple[dict, dict]]:
    if sample_size > len(records):
        raise ValueError("사람 검토 표본이 검수 JSONL 건수보다 클 수 없습니다.")
    selected: list[tuple[dict, dict]] = []
    selected_ids: set[str] = set()
    counts = {
        "source_channel": Counter(),
        "language": Counter(),
        "generation_profile_id": Counter(),
    }

    def add(record: tuple[dict, dict]) -> None:
        document, generation = record
        if document["voc_id"] in selected_ids:
            return
        selected.append(record)
        selected_ids.add(document["voc_id"])
        counts["source_channel"][document["source_channel"]] += 1
        counts["language"][document["language"]] += 1
        counts["generation_profile_id"][generation["generation_profile_id"]] += 1

    # Every S4 document in the review source is mandatory.
    for record in records:
        document, _ = record
        if any(issue["severity"] == "S4" for issue in document["issues"]):
            add(record)

    dimension_values = {
        "source_channel": sorted({doc["source_channel"] for doc, _ in records}),
        "language": sorted({doc["language"] for doc, _ in records}),
        "generation_profile_id": sorted(
            {meta["generation_profile_id"] for _, meta in records}
        ),
    }
    while len(selected) < sample_size:
        best = None
        best_score = -1
        for record in records:
            document, generation = record
            if document["voc_id"] in selected_ids:
                continue
            score = 0
            values = {
                "source_channel": document["source_channel"],
                "language": document["language"],
                "generation_profile_id": generation["generation_profile_id"],
            }
            for dimension, value in values.items():
                deficit = max(0, min_per_stratum - counts[dimension][value])
                score += deficit * 10
            parent_scenarios = [
                scenario_by_id[parent_id]
                for parent_id in document["synthetic_parent_scenario_ids"]
            ]
            if len(document["issues"]) > 1:
                score += 5
            if any(scenario["hard_negative"] == "TRUE" for scenario in parent_scenarios):
                score += 4
            if any(issue["safety_flags"] != ["NONE"] for issue in document["issues"]):
                score += 3
            if score > best_score:
                best = record
                best_score = score
        if best is None:
            break
        add(best)

    return selected[:sample_size]


def build_human_review_workbook(
    review_path: Path,
    source_path: Path,
    output_path: Path,
    sample_size: int = 250,
    min_per_stratum: int = 20,
) -> dict:
    review = json.loads(review_path.read_text(encoding="utf-8"))
    sample = review["sample"]
    documents = _read_jsonl(review_path.parent / sample["data_file"])
    generations = _read_jsonl(review_path.parent / sample["generation_file"])
    if len(documents) != len(generations):
        raise ValueError("sample 본문과 generation sidecar 건수가 다릅니다.")
    scenarios = load_scenarios(source_path)
    scenario_by_id = {scenario.scenario_id: scenario for scenario in scenarios}
    selected = _select_review_rows(
        list(zip(documents, generations)),
        scenario_by_id,
        sample_size,
        min_per_stratum,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "human_review"
    headers = [
        "voc_id",
        "dataset_split",
        "language",
        "source_channel",
        "generation_profile_id",
        "parent_scenario_ids",
        "is_multi_issue",
        "hard_negative",
        "severity",
        "safety_flags",
        "raw_text",
        "issue_labels",
        "label_match",
        "naturalness_1_5",
        "channel_fit_1_5",
        "cause_overclaim",
        "reviewer_notes",
    ]
    sheet.append(headers)
    for document, generation in selected:
        parents = document["synthetic_parent_scenario_ids"]
        parent_rows = [scenario_by_id[parent_id] for parent_id in parents]
        issue_labels = " | ".join(
            f"{issue['affected_function']}::{issue['observed_symptom']}"
            for issue in document["issues"]
        )
        sheet.append(
            [
                document["voc_id"],
                document["dataset_split"],
                document["language"],
                document["source_channel"],
                generation["generation_profile_id"],
                "|".join(parents),
                "YES" if len(document["issues"]) > 1 else "NO",
                "YES" if any(row["hard_negative"] == "TRUE" for row in parent_rows) else "NO",
                "|".join(issue["severity"] for issue in document["issues"]),
                "|".join(
                    flag
                    for issue in document["issues"]
                    for flag in issue["safety_flags"]
                ),
                document["raw_text"],
                issue_labels,
                "",
                "",
                "",
                "",
                "",
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 24,
        "B": 12,
        "C": 14,
        "D": 26,
        "E": 22,
        "F": 28,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 24,
        "K": 80,
        "L": 55,
        "M": 14,
        "N": 18,
        "O": 18,
        "P": 18,
        "Q": 40,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    yes_no = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    score = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="5",
        allow_blank=True,
    )
    sheet.add_data_validation(yes_no)
    sheet.add_data_validation(score)
    yes_no.add(f"M2:M{sheet.max_row}")
    yes_no.add(f"P2:P{sheet.max_row}")
    score.add(f"N2:O{sheet.max_row}")
    sheet.conditional_formatting.add(
        f"N2:O{sheet.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=1,
            start_color="F8696B",
            mid_type="num",
            mid_value=3,
            mid_color="FFEB84",
            end_type="num",
            end_value=5,
            end_color="63BE7B",
        ),
    )

    summary = workbook.create_sheet("sampling_summary")
    summary.append(["review_id", review["review_id"]])
    summary.append(["source_rows", len(documents)])
    summary.append(["selected_rows", len(selected)])
    summary.append(["min_per_stratum_target", min_per_stratum])
    summary.append([])
    summary.append(["dimension", "value", "count"])
    for dimension, getter in (
        ("language", lambda d, g: d["language"]),
        ("source_channel", lambda d, g: d["source_channel"]),
        ("generation_profile_id", lambda d, g: g["generation_profile_id"]),
        ("issue_count", lambda d, g: str(len(d["issues"]))),
    ):
        counts = Counter(getter(document, generation) for document, generation in selected)
        for value, count in sorted(counts.items()):
            summary.append([dimension, value, count])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "workbook": output_path,
        "source_review": review_path,
        "source_rows": len(documents),
        "selected_rows": len(selected),
    }
